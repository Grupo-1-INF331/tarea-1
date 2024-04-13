from openpyxl import load_workbook
import random
import hashlib
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import padding
import os

def hash_password(password):
    # Create a new SHA-256 hash object
    hash_object = hashlib.sha256()
    
    # Convert the password string to bytes, as the hashing function requires byte input
    password_bytes = password.encode('utf-8')
    
    # Update the hash object with the bytes of the password
    hash_object.update(password_bytes)
    
    # Get the hexadecimal representation of the hash
    hashed_password = hash_object.hexdigest()
    
    # Convert hex digest to bytes to use as encryption key
    key_bytes = bytes.fromhex(hashed_password)
    
    return key_bytes

def confirm_username(username):
    wb = load_workbook('users_data.xlsx')
    hoja = wb.worksheets[0]
    repetido = False
    for fila in hoja.iter_rows(min_row = 2):
        val_row = [celda.value for celda in fila]
        print(val_row[0])
        print(username)
        if username == val_row[0] :
            repetido = True
    return repetido

def check_password(password, confirmpassword):
    if len(password) < 8:
        print("Password must be at least 8 characters long.")
        return False
    
    # Check if password and confirm password match
    if password != confirmpassword:
        print("Passwords do not match. Please try again.")
        return False
    
    # Check if password contains only allowed characters
    allowed_chars = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz°.!\"#$%&/()<=¿?¡{}*_+-[].@[^;"
    for char in password:
        if char not in allowed_chars:
            print("Password contains invalid characters. Please try again.")
            return False
    return True

def confirm_password(username, password):
    wb = load_workbook('users_data.xlsx')
    hoja = wb.worksheets[0]
    for fila in hoja.iter_rows(min_row = 2):
        val_row = [celda.value for celda in fila]
        if val_row[0] == username :
            user_db = wb.worksheets[int(val_row[2])]
            for row in user_db.iter_rows(min_row=2, max_row=2):
                val_fila = [celda.value for celda in row]
                hashed_password = hash_password(password)
                if hash_password == val_fila[1]: 
                    return True
    print("Contraseña errónea!\n")
    return False

def generate_password():
    # Define the default characters
    numbers = '0123456789'
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'
    special_chars = '°.!\"#$%&/()<=>?¡{}*_+-.[\]@[;^'
    
    # Prompt user for the password length
    length = int(input("Ingrese la longitud de contraseña deseada (mínimo 8): "))
    while length < 8:
        length = int(input("Longitud demasiado corta. Por favor ingrese un mínimo de 8: "))
    
    # Allow users to add or remove characters
    print("Se utiliza el juego de caracteres predeterminado. ¿Quieres personalizarlo? (si/no): ")
    customize = input().strip().lower()
    
    allowed_characters = numbers + letters + special_chars
    if customize == 'si':
        print("Ingrese su conjunto personalizado de caracteres: ")
        allowed_characters = input()
    
    # Generate the password
    password = ''.join(random.choice(allowed_characters) for _ in range(length))
    return password

# Function to encrypt a password
def encrypt_password(password, key):
    # Ensure the key is 32 bytes for AES-256
    if len(key) != 32:
        raise ValueError("La clave debe tener 32 bytes de longitud.")
    
    # Setup cipher configuration
    backend = default_backend()
    iv = os.urandom(16)  # Generate a random 16-byte IV.
    cipher = Cipher(algorithms.AES(key), modes.CFB(iv), backend=backend)
    encryptor = cipher.encryptor()
    
    # Pad the password to ensure it's a multiple of block size
    padder = padding.PKCS7(algorithms.AES.block_size).padder()
    padded_data = padder.update(password.encode()) + padder.finalize()
    
    # Encrypt the password
    encrypted = encryptor.update(padded_data) + encryptor.finalize()
    
    # Return IV + encrypted password (IV is needed for decryption)
    return iv + encrypted

# Function to decrypt a password
def decrypt_password(encrypted_password, key):
    # Ensure the key is 32 bytes for AES-256
    if len(key) != 32:
        raise ValueError("La clave debe tener 32 bytes de longitud.")
    
    # Extract the IV (first 16 bytes)
    iv = encrypted_password[:16]
    encrypted_password = encrypted_password[16:]
    
    # Setup cipher configuration
    backend = default_backend()
    cipher = Cipher(algorithms.AES(key), modes.CFB(iv), backend=backend)
    decryptor = cipher.decryptor()
    
    # Decrypt the password
    decrypted_padded = decryptor.update(encrypted_password) + decryptor.finalize()
    
    # Unpad the decrypted password
    unpadder = padding.PKCS7(algorithms.AES.block_size).unpadder()
    decrypted = unpadder.update(decrypted_padded) + unpadder.finalize()
    
    return decrypted.decode()
